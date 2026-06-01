import io
import os
from openpyxl import load_workbook


def _get_writeable_cell(ws, cell_ref: str) -> str:
    """
    Dacă cell_ref este într-un range merged, întoarce celula top-left
    a acelui merged range. Altfel întoarce cell_ref nemodificat.
    """
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            return merged_range.start_cell.coordinate
    return cell_ref


def render_excel_template(template_path: str, cell_values: dict, sheet_name: str | None = None) -> bytes:
    """
    Încarcă un template Excel, completează celulele indicate și returnează
    fișierul rezultat ca bytes.

    - template_path: calea către fișierul .xlsx
    - cell_values: dict de forma { "A1": "valoare", "D20": 2550.0, ... }
    - sheet_name: numele sheet-ului; dacă lipsește, folosește ws active
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Nu găsesc template-ul Excel: {template_path}")

    wb = load_workbook(template_path)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet-ul '{sheet_name}' nu există în template.")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    for cell_ref, value in cell_values.items():
        target_cell = _get_writeable_cell(ws, cell_ref)
        ws[target_cell] = value

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()